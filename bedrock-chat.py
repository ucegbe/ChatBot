import boto3
from botocore.config import Config
import shutil
import os
import pandas as pd
import json
import io
import streamlit as st
import time
config = Config(
    read_timeout=600,
    retries = dict(
        max_attempts = 5 ## Handle retries
    )
)
import re
import numpy as np
import openpyxl
import base64
from openpyxl.cell import Cell
from openpyxl.worksheet.cell_range import CellRange
from textractor import Textractor
from textractor.visualizers.entitylist import EntityList
from textractor.data.constants import TextractFeatures
from textractor.data.text_linearization_config import TextLinearizationConfig
from boto3.dynamodb.conditions import Key 
import time
import random
import logging
from botocore.exceptions import ClientError
# Read credentials
with open('config.json') as f:
    config_file = json.load(f)
# pricing info
with open('pricing.json') as f:
    pricing_file = json.load(f)

DYNAMODB_TABLE=config_file["DynamodbTable"]
BUCKET=config_file["Bucket_Name"]
OUTPUT_TOKEN=config_file["max-output-token"]
S3 = boto3.client('s3')
DYNAMODB  = boto3.resource('dynamodb')
st.set_page_config(initial_sidebar_state="auto")
COGNITO = boto3.client('cognito-idp')
S3_DOC_CACHE_PATH='uploads'
TEXTRACT_RESULT_CACHE_PATH="textract_output"
LOAD_DOC_IN_ALL_CHAT_CONVO=config_file["load-doc-in-chat-history"]
CHAT_HISTORY_LENGTH=config_file["chat-history-loaded-length"]
DYNAMODB_USER=config_file["UserId"]
APP_ID=config_file["cognito-app-id"]
USE_COGNITO=config_file["use-cognito"]
REGION=config_file["bedrock-region"]
bedrock_runtime = boto3.client(service_name='bedrock-runtime',region_name=REGION,config=config)

if 'messages' not in st.session_state:
    st.session_state['messages'] = []
if 'input_token' not in st.session_state:
    st.session_state['input_token'] = 0
if 'output_token' not in st.session_state:
    st.session_state['output_token'] = 0
if 'chat_hist' not in st.session_state:
    st.session_state['chat_hist'] = []
if 'user_sess' not in st.session_state:
    st.session_state['user_sess'] =f"{DYNAMODB_USER}-{str(time.time()).split('.')[0]}"
if 'chat_session_list' not in st.session_state:
    st.session_state['chat_session_list'] = []
if 'count' not in st.session_state:
    st.session_state['count'] = 0
if 'userid' not in st.session_state:
    st.session_state['userid']= config_file["UserId"]
if 'cost' not in st.session_state:
    st.session_state['cost'] = 0

def handle_doc_upload_or_s3(file):
    """
    Handle parsing of documents uploaded.
    Supports PDF, PNG, JPG, CSV, XLSX, JSON, Python Scripts and Text files.
    Parameters:
        file (str): S3 URI of document to parse
    Returns:
        content: Parsed contents of the file in appropriate format
    """    
    dir_name, ext = os.path.splitext(file)
    if  ext in [".pdf", ".png", ".jpg"]:   
        content=exract_pdf_text_aws(file)
    elif "csv"  in ext:
        content= pd.read_csv(file)
    elif ext in [".xlsx", ".xlx"]:
        content=table_parser_utills(file)   
    elif  "json" in ext:      
        obj=get_s3_obj_from_bucket_(file)
        content = json.loads(obj['Body'].read())  
    elif  ext in [".txt",".py"]:       
        obj=get_s3_obj_from_bucket_(file)
        content = obj['Body'].read()
    # Implement any of file extension logic 
    return content


def exract_pdf_text_aws(file):
    """
    Extract text from PDF/image files using Amazon Textract service.
    Supports PDFs/Images stored uploaded.    
    Parameters:
        file (str):S3 URI of PDF file
    Returns:
        text (str): Extracted text from PDF
    """     
    file_base_name=os.path.basename(file)
    if [x for x in get_s3_keys(f"{TEXTRACT_RESULT_CACHE_PATH}/") if file_base_name in x]:      
        response = S3.get_object(Bucket=BUCKET, Key=f"{TEXTRACT_RESULT_CACHE_PATH}/{file_base_name}.txt")
        text = response['Body'].read()
        return text
    
    else:
        dir_name, ext = os.path.splitext(file)
        extractor = Textractor(region_name="us-east-1")
        # Asynchronous call, you will experience some wait time. Try caching results for better experience
        if "pdf" in ext:
            logga=st.empty()
            logga.write("Asynchronous Textract call, you may experience some wait time.")
            document = extractor.start_document_analysis(
            file_source=file,
            features=[TextractFeatures.LAYOUT,TextractFeatures.TABLES],       
            save_image=False,   
            s3_output_path=f"s3://{BUCKET}/textract_output/" #S3 location to store textract json response
        )    
            logga.write("")
        #Synchronous call
        else:
            document = extractor.analyze_document(
            file_source=file,
            features=[TextractFeatures.LAYOUT,TextractFeatures.TABLES],  
            save_image=False,
        )

        config = TextLinearizationConfig(
        hide_figure_layout=True,   
        hide_header_layout=False,    
        table_prefix="<table>",
        table_suffix="</table>",
        )      
        S3.put_object(Body=document.get_text(config=config), Bucket=BUCKET, Key=f"{TEXTRACT_RESULT_CACHE_PATH}/{file_base_name}.txt") 
        return document.get_text(config=config)

def strip_newline(cell):
    """
    A utility function to strip newline characters from a cell.
    Parameters:
    cell (str): The cell value.
    Returns:
    str: The cell value with newline characters removed.
    """
    return str(cell).strip()

def table_parser_utills(file):    
    """
    Converts an Excel table to a csv string, 
    handling duplicated values across merged cells.
    Args:
        file: Excel table  
    Returns: 
        Pandas DataFrame representation of the Excel table
    """
    # Read from S3 or local    
    s3 = boto3.client('s3')
    match = re.match("s3://(.+?)/(.+)", file)
    if match:
        bucket_name = match.group(1)
        key = match.group(2)    
        obj = s3.get_object(Bucket=bucket_name, Key=key)  
    # Read Excel file from S3 into a buffer
    xlsx_buffer = io.BytesIO(obj['Body'].read())
    xlsx_buffer.seek(0) 
    # Load workbook, get active worksheet
    wb = openpyxl.load_workbook(xlsx_buffer)
    worksheet = wb.active
    # Unmerge cells, duplicate merged values to individual cells
    all_merged_cell_ranges: list[CellRange] = list(
            worksheet.merged_cells.ranges
        )
    for merged_cell_range in all_merged_cell_ranges:
        merged_cell: Cell = merged_cell_range.start_cell
        worksheet.unmerge_cells(range_string=merged_cell_range.coord)
        for row_index, col_index in merged_cell_range.cells:
            cell: Cell = worksheet.cell(row=row_index, column=col_index)
            cell.value = merged_cell.value
    # determine table header index
    df = pd.DataFrame(worksheet.values)
    df=df.map(strip_newline)  
    return df.to_csv(sep="|", index=False, header=0)


def put_db(params,messages):
    """Store long term chat history in DynamoDB"""    
    chat_item = {
        "UserId": st.session_state['userid'], # user id
        "SessionId": params["session_id"], # User session id
        "messages": [messages],  # 'messages' is a list of dictionaries
        "time":messages['time']
    }

    existing_item = DYNAMODB.Table(DYNAMODB_TABLE).get_item(Key={"UserId": st.session_state['userid'], "SessionId":params["session_id"]})
    if "Item" in existing_item:
        existing_messages = existing_item["Item"]["messages"]
        chat_item["messages"] = existing_messages + [messages]
    response = DYNAMODB.Table(DYNAMODB_TABLE).put_item(
        Item=chat_item
    )
    
def get_chat_history_db(params,cutoff):
    current_chat, chat_hist=[],[]
    claude3=False
    if "sonnet" in params['model'] or "haiku" in params['model']:
        claude3=True
    if "Item" in params['chat_histories']:  
        chat_hist=params['chat_histories']['Item']['messages'][-cutoff:]            
        for d in chat_hist:
            if d['image'] and claude3 and LOAD_DOC_IN_ALL_CHAT_CONVO:
                content=[]
                for img in d['image']:
                    s3 = boto3.client('s3')
                    match = re.match("s3://(.+?)/(.+)", img)
                    image_name=os.path.basename(img)
                    _,ext=os.path.splitext(image_name)
                    if "jpg" in ext: ext=".jpeg"                        
                    if match:
                        bucket_name = match.group(1)
                        key = match.group(2)    
                        obj = s3.get_object(Bucket=bucket_name, Key=key)
                        base_64_encoded_data = base64.b64encode(obj['Body'].read())
                        base64_string = base_64_encoded_data.decode('utf-8')                        
                    content.extend([{"type":"text","text":image_name},{
                      "type": "image",
                      "source": {
                        "type": "base64",
                        "media_type": f"image/{ext.lower().replace('.','')}",
                        "data": base64_string
                      }
                    }])
                content.extend([{"type":"text","text":d['user']}])
                current_chat.append({'role': 'user', 'content': content})
            elif d['document'] and LOAD_DOC_IN_ALL_CHAT_CONVO:
                doc='Here are the documents:\n'
                for docs in d['document']:
                    uploads=handle_doc_upload_or_s3(docs)
                    doc_name=os.path.basename(docs)
                    doc+=f"<{doc_name}>\n{uploads}\n</{doc_name}>\n"
                if not claude3 and d["image"]:
                    for docs in d['image']:
                        uploads=handle_doc_upload_or_s3(docs)
                        doc_name=os.path.basename(docs)
                        doc+=f"<{doc_name}>\n{uploads}\n</{doc_name}>\n"
                current_chat.append({'role': 'user', 'content': [{"type":"text","text":doc+d['user']}]})
            else:
                current_chat.append({'role': 'user', 'content': [{"type":"text","text":d['user']}]})
            current_chat.append({'role': 'assistant', 'content': d['assistant']})  
    else:
        chat_hist=[]
    return current_chat, chat_hist

    
def get_s3_keys(prefix):
    s3 = boto3.client('s3')
    response = s3.list_objects_v2(Bucket=BUCKET, Prefix=prefix)
    keys=""
    if "Contents" in response:
        keys = []
        for obj in response['Contents']:
            key = obj['Key']
            name = key[len(prefix):]
            keys.append(name)
    return keys

def get_s3_obj_from_bucket_(file):
    """Retrieves an object from an S3 bucket given its S3 URI.
    Args:
       file (str): The S3 URI of the object to retrieve, in the format "s3://{bucket_name}/{key}".
   Returns:
       botocore.response.StreamingBody: The retrieved S3 object.
    """
    s3 = boto3.client('s3')
    match = re.match("s3://(.+?)/(.+)", file)
    if match:
        bucket_name = match.group(1)
        key = match.group(2)    
        obj = s3.get_object(Bucket=bucket_name, Key=key)  
    return obj

def put_obj_in_s3_bucket_(docs):
    """Uploads a file to an S3 bucket and returns the S3 URI of the uploaded object.
    Args:
       docs (str): The local file path of the file to upload to S3.
   Returns:
       str: The S3 URI of the uploaded object, in the format "s3://{bucket_name}/{file_path}".
    """
    file_name=os.path.basename(docs.name)
    file_path=f"{S3_DOC_CACHE_PATH}/{file_name}"
    S3.put_object(Body=docs.read(),Bucket= BUCKET, Key=file_path)
    return f"s3://{BUCKET}/{file_path}"


def bedrock_streemer(params,response, handler):
    stream = response.get('body')
    answer = ""    
    if stream:
        for event in stream:
            chunk = event.get('chunk')
            if  chunk:
                chunk_obj = json.loads(chunk.get('bytes').decode())
                if "delta" in chunk_obj:                    
                    delta = chunk_obj['delta']
                    if "text" in delta:
                        text=delta['text'] 
                        # st.write(text, end="")                        
                        answer+=str(text)       
                        handler.markdown(answer.replace("$","USD ").replace("%", " percent"))
                        
                if "amazon-bedrock-invocationMetrics" in chunk_obj:
                    st.session_state['input_token'] = chunk_obj['amazon-bedrock-invocationMetrics']['inputTokenCount']
                    st.session_state['output_token'] =chunk_obj['amazon-bedrock-invocationMetrics']['outputTokenCount']
                    pricing=st.session_state['input_token']*pricing_file[f"anthropic.{params['model']}"]["input"]+st.session_state['output_token'] *pricing_file[f"anthropic.{params['model']}"]["output"]
                    st.session_state['cost']+=pricing             
    return answer

def bedrock_claude_(params,chat_history,system_message, prompt,model_id,image_path=None, handler=None):
    # content=[{
    #     "type": "text",
    #     "text": prompt
    #         }]
    content=[]
    if image_path:       
        if not isinstance(image_path, list):
            image_path=[image_path]      
        for img in image_path:
            s3 = boto3.client('s3')
            match = re.match("s3://(.+?)/(.+)", img)
            image_name=os.path.basename(img)
            _,ext=os.path.splitext(image_name)
            if "jpg" in ext: ext=".jpeg"                        
            if match:
                bucket_name = match.group(1)
                key = match.group(2)    
                obj = s3.get_object(Bucket=bucket_name, Key=key)
                base_64_encoded_data = base64.b64encode(obj['Body'].read())
                base64_string = base_64_encoded_data.decode('utf-8')
            content.extend([{"type":"text","text":image_name},{
              "type": "image",
              "source": {
                "type": "base64",
                "media_type": f"image/{ext.lower().replace('.','')}",
                "data": base64_string
              }
            }])
    content.append({
        "type": "text",
        "text": prompt
            })
    chat_history.append({"role": "user",
            "content": content})
    # print(system_message)
    prompt = {
        "anthropic_version": "bedrock-2023-05-31",
        "max_tokens": 1500,
        "temperature": 0.5,
        "system":system_message,
        "messages": chat_history
    }
    answer = ""
    prompt = json.dumps(prompt)
    response = bedrock_runtime.invoke_model_with_response_stream(body=prompt, modelId=model_id, accept="application/json", contentType="application/json")
    answer=bedrock_streemer(params,response, handler) 
    return answer

def _invoke_bedrock_with_retries(params,current_chat, chat_template, question, model_id, image_path, handler):
    max_retries = 10
    backoff_base = 2
    max_backoff = 3  # Maximum backoff time in seconds
    retries = 0

    while True:
        try:
            response = bedrock_claude_(params,current_chat, chat_template, question, model_id, image_path, handler)
            return response
        except ClientError as e:
            if e.response['Error']['Code'] == 'ThrottlingException':
                if retries < max_retries:
                    # Throttling, exponential backoff
                    sleep_time = min(max_backoff, backoff_base ** retries + random.uniform(0, 1))
                    time.sleep(sleep_time)
                    retries += 1
                else:
                    raise e
            else:
                # Some other API error, rethrow
                raise
                

def get_session_ids_by_user(table_name, user_id):
    """
    Get Session Ids and corresponding top message for a user to populate the chat history drop down on the front end
    """
    table = DYNAMODB.Table(table_name)
    message_list={}
    session_ids = []
    args = {
        'KeyConditionExpression': Key('UserId').eq(user_id)
    }
    while True:
        response = table.query(**args)
        session_ids.extend([item['SessionId'] for item in response['Items']])
        if 'LastEvaluatedKey' not in response:
            break
        args['ExclusiveStartKey'] = response['LastEvaluatedKey']
        
    for session_id in session_ids:
        try:
            message_list[session_id]=DYNAMODB.Table(table_name).get_item(Key={"UserId": user_id, "SessionId":session_id})['Item']['messages'][0]['user']
        except Exception as e:
            print(e)
            pass
    return message_list


def query_llm(params, handler):
    """
    Function takes a user query and a uploaded document. Caches documents in S3
    passing a document is optional
    """  
    if not isinstance(params['upload_doc'], list):
        raise TypeError("documents must be in a list format")        
    # Check if Claude3 model is used and handle images with the CLAUDE3 Model
    claude3=False
    model='anthropic.'+params['model']
    if "sonnet" in model or "haiku" in model:
        model+="-20240229-v1:0" if "sonnet" in model else "-20240307-v1:0"
        claude3=True
    # Retrieve past chat history from Dynamodb
    if DYNAMODB_TABLE:
        current_chat,chat_hist=get_chat_history_db(params, CHAT_HISTORY_LENGTH)
        # st.write(chat_hist)
        # st.stop()
    else:
        for d in chat_hist:
            current_chat.append({'role': 'user', 'content': d['user']})
            current_chat.append({'role': 'assistant', 'content': d['assistant']})
    # print(current_chat)
    ## prompt template for when a user uploads a doc
    doc_path=[]
    image_path=[]
    doc=""
    if params['upload_doc']:  
        doc='Here are documents:\n'
        for ids,docs in enumerate(params['upload_doc']):
            file_name=docs.name
            _,extensions=os.path.splitext(file_name)
            docs=put_obj_in_s3_bucket_(docs)
            if extensions in [".jpg",".jpeg",".png",".gif",".webp"] and claude3:
                print("yes")
                image_path.append(docs)
                continue
            uploads=handle_doc_upload_or_s3(docs)             
            doc_path.append(docs)
            
            doc_name=os.path.basename(docs)
            doc+=f"<{doc_name}>\n{uploads}\n</{doc_name}>\n"
        with open("prompt/doc_chat.txt","r") as f:
            chat_template=f.read() 
        
        
        document_values = []
        for items in chat_hist:
            document_values.extend(items["document"])
        difference = set(document_values) ^ set(doc_path)   
        if not difference:
            doc=""
            doc_path=[]
        
    else:        
        # Chat template for open ended query
        with open("prompt/chat.txt","r") as f:
            chat_template=f.read()

    response=_invoke_bedrock_with_retries(params,current_chat, chat_template, doc+params['question'], model, image_path, handler)
    # log the following items to dynamodb
    chat_history={"user":params['question'],
    "assistant":response,
    "image":image_path,
    "document":doc_path,
    "modelID":model,
    "time":str(time.time()),
    "input_token":round(st.session_state['input_token']) ,
    "output_token":round(st.session_state['output_token'])} 
    #store convsation memory and user other items in DynamoDB table
    if DYNAMODB_TABLE:
        put_db(params,chat_history)
    # use local memory for storage
    else:
        chat_hist.append(chat_history)   
    return response


def get_chat_historie_for_streamlit(params):
    """
    This function retrieves chat history stored in a dynamoDB table partitioned by a userID and sorted by a SessionID
    """
    chat_histories = DYNAMODB.Table(DYNAMODB_TABLE).get_item(Key={"UserId": st.session_state['userid'], "SessionId":params["session_id"]})

# Constructing the desired list of dictionaries
    formatted_data = []
    if 'Item' in chat_histories:
        for entry in chat_histories['Item']['messages']:
            # assistant_attachment = entry.get('image', []) + entry.get('document', [])
            assistant_attachment = '\n\n'.join(entry.get('image', []) + entry.get('document', []))
            formatted_data.append({
                "role": "user",
                "content": entry["user"],
            })
            formatted_data.append({
                "role": "assistant",
                "content": entry["assistant"],
                "attachment": assistant_attachment
            })
    else:
        chat_histories=[]            
    return formatted_data,chat_histories



def get_key_from_value(dictionary, value):
    return next((key for key, val in dictionary.items() if val == value), None)
    
def chat_bedrock_(params):
    st.title('Chatty AI Assitant 🙂')
    params['chat_histories']=[]
    if params["session_id"].strip():
        st.session_state.messages, params['chat_histories']=get_chat_historie_for_streamlit(params)
    for message in st.session_state.messages:
        with st.chat_message(message["role"]):   
            if "```" in message["content"]:
                st.markdown(message["content"],unsafe_allow_html=True )
            else:
                st.markdown(message["content"].replace("$", "\$"),unsafe_allow_html=True )
            if message["role"]=="assistant":
                if message["attachment"]:
                    with st.expander(label="**attachments**"):
                        st.markdown(message["attachment"])
    if prompt := st.chat_input("Whats up?"):        
        st.session_state.messages.append({"role": "user", "content": prompt})        
        with st.chat_message("user"):             
            if "```" in prompt:
                st.markdown(prompt,unsafe_allow_html=True )
            else:
                st.markdown(prompt.replace("$", "\$"),unsafe_allow_html=True )
        with st.chat_message("assistant"): 
            message_placeholder = st.empty()
            time_now=time.time()            
            params["question"]=prompt
            answer=query_llm(params, message_placeholder)
            if "```" in answer:
                message_placeholder.markdown(answer,unsafe_allow_html=True )
            else:
                message_placeholder.markdown(answer.replace("$", "\$"),unsafe_allow_html=True )
            st.session_state.messages.append({"role": "assistant", "content": answer}) 
        st.rerun()
        
def app_sidebar():
    with st.sidebar:   
        st.metric(label="Bedrock Session Cost", value=f"${round(st.session_state['cost'],2)}") 
        st.write("-----")
        button=st.button("New Chat", type ="primary")
        models=[ 'claude-3-sonnet','claude-3-haiku','claude-instant-v1','claude-v2:1', 'claude-v2']
        model=st.selectbox('**Model**', models)
        params={"model":model}       
        user_sess_id=get_session_ids_by_user(DYNAMODB_TABLE, st.session_state['userid'])
        dict_items = list(user_sess_id.items())
        dict_items.insert(0, (st.session_state['user_sess'],"New Chat")) 
        if button:
            st.session_state['user_sess'] = f"{st.session_state['userid']}-{str(time.time()).split('.')[0]}"
            dict_items.insert(0, (st.session_state['user_sess'],"New Chat"))      
        st.session_state['chat_session_list'] = dict(dict_items)
        chat_items=st.selectbox("**Chat Sessions**",st.session_state['chat_session_list'].values(),key="chat_sessions")
        session_id=get_key_from_value(st.session_state['chat_session_list'], chat_items)   
        file = st.file_uploader('Upload a document', accept_multiple_files=True, help="pdf,csv,txt,png,jpg,xlsx,json,py doc format supported") 
        params={"model":model, "session_id":session_id, "chat_item":chat_items, "upload_doc":file }    
        st.session_state['count']=1
        return params

def check_password():
    """Returns `True` if the user had a correct password."""
    def login_form():
        """Form with widgets to collect user information"""
        with st.form("Credentials"):
            st.text_input("Username", key="username")
            st.text_input("Password", type="password", key="password")
            st.form_submit_button("Log in", on_click=password_entered)
    def password_entered():
        """Checks whether a password entered by the user is correct."""
        try:
            cred=COGNITO.initiate_auth(
                AuthFlow='USER_PASSWORD_AUTH',
                AuthParameters={
                 "USERNAME": st.session_state["username"],
                "PASSWORD": st.session_state["password"],
                },
                ClientId=APP_ID,
            )              
            st.session_state["password_correct"] = True
            del st.session_state["password"]  # Don't store the username or password.  
            st.session_state['userid']=st.session_state['username']
            st.session_state['user_sess'] = f"{st.session_state['username']}-{str(time.time()).split('.')[0]}"   
            del st.session_state["username"]      
        except COGNITO.exceptions.NotAuthorizedException as e:
            if "Incorrect username or password" in str(e):
                st.session_state["password_correct"] = False
        except client.exceptions.PasswordResetRequiredException:
            st.error("Password reset is required. Please reset your password.")
        except ClientError as e:
            st.error(f"An error occurred: {str(e)}")
        except Exception as e:
            st.error(f"An unexpected error occurred: {str(e)}")

    if st.session_state.get("password_correct", False):
        return True
    login_form()
    if "password_correct" in st.session_state:
        st.error("😕 Username or password incorrect")
    return False
    
def main():    
    if USE_COGNITO:
        if not check_password():
            st.stop()
    params=app_sidebar()
    chat_bedrock_(params)
   
    
if __name__ == '__main__':
    main()   
    
